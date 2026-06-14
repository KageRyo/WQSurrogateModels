from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--output-file", default=None)
    return parser.parse_args()


def read_csv_if_exists(path: Path) -> pd.DataFrame:
    return pd.read_csv(path) if path.exists() and path.stat().st_size > 0 else pd.DataFrame()


def environment_frame(output_dir: Path) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for section, filename in [("manifest", "manifest.json"), ("hardware", "hardware.json"), ("versions", "versions.json")]:
        path = output_dir / filename
        if not path.exists():
            continue
        data = json.loads(path.read_text(encoding="utf-8"))
        for key, value in data.items():
            rows.append(
                {
                    "section": section,
                    "key": key,
                    "value": json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value,
                }
            )
    return pd.DataFrame(rows)


def file_index(output_dir: Path) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for path in sorted(output_dir.rglob("*")):
        if path.is_file():
            rows.append(
                {
                    "relative_path": str(path.relative_to(output_dir)),
                    "size_mb": round(path.stat().st_size / (1024 * 1024), 3),
                    "modified_time": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
                }
            )
    return pd.DataFrame(rows)


def prediction_counts(output_dir: Path) -> pd.DataFrame:
    path = output_dir / "predictions" / "predictions_long.csv"
    if not path.exists():
        return pd.DataFrame()
    return (
        pd.read_csv(path, usecols=["source", "missing_set", "experiment_mode", "experiment", "seed", "model_type"])
        .groupby(["source", "missing_set", "experiment_mode", "experiment", "seed", "model_type"], as_index=False)
        .size()
        .rename(columns={"size": "prediction_rows"})
    )


def add_sheet(wb: Workbook, name: str, frame: pd.DataFrame) -> None:
    ws = wb.create_sheet(name[:31])
    if frame.empty:
        ws.append(["No rows"])
        return
    ws.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        ws.append(list(row))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row >= 2 and max_col >= 1:
        table_name = ("T_" + "".join(ch if ch.isalnum() else "_" for ch in name))[:30]
        table = Table(displayName=table_name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    numeric_suffixes = ("_mean", "_std", "_s", "_ms", "_value", "_low", "_high", "mae", "rmse", "r2")
    for col_idx, header in enumerate(frame.columns, start=1):
        letter = get_column_letter(col_idx)
        width = min(max(len(str(header)) + 2, 10), 45)
        for row_idx in range(2, min(max_row, 300) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                width = min(max(width, len(str(value)) + 2), 45)
        ws.column_dimensions[letter].width = width
        if str(header).endswith(numeric_suffixes) or str(header) in {"r2", "mae", "rmse", "accuracy", "macro_f1"}:
            for row_idx in range(2, max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = "0.0000"
    for metric in ["r2_mean", "accuracy_mean", "macro_f1_mean"]:
        if metric in frame.columns and max_row >= 3:
            col = get_column_letter(list(frame.columns).index(metric) + 1)
            ws.conditional_formatting.add(
                f"{col}2:{col}{max_row}",
                ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
            )
    for metric in ["mae_mean", "rmse_mean", "nmae_mean"]:
        if metric in frame.columns and max_row >= 3:
            col = get_column_letter(list(frame.columns).index(metric) + 1)
            ws.conditional_formatting.add(
                f"{col}2:{col}{max_row}",
                ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"),
            )


def main() -> None:
    args = parse_args()
    output_dir = (PROJECT_ROOT / args.output_dir).resolve()
    output_file = Path(args.output_file) if args.output_file else output_dir / "reports" / "missing_indicator_robustness_summary.xlsx"
    if not output_file.is_absolute():
        output_file = PROJECT_ROOT / output_file
    output_file.parent.mkdir(parents=True, exist_ok=True)

    readme = pd.DataFrame(
        [
            {"item": "Workbook purpose", "value": "Summary of missing-indicator robustness experiments."},
            {"item": "Result directory", "value": str(output_dir)},
            {"item": "Row-level predictions", "value": "See predictions/predictions_long.csv; Excel contains grouped counts and summaries."},
            {"item": "Stress test", "value": "event-window stress is controlled synthetic perturbation, not real event validation."},
            {"item": "Stress107", "value": "107 sequential event-window stress sheets are included when stress107 CSV files exist."},
            {"item": "CPU timing", "value": "CPU-only timing is included if timing/cpu_only_inference_timing_summary.csv exists."},
        ]
    )

    sheet_specs = [
        ("README", readme),
        ("Best_Summary", read_csv_if_exists(output_dir / "metrics" / "best_by_experiment_source.csv")),
        ("Metrics_Summary", read_csv_if_exists(output_dir / "metrics" / "metrics_summary.csv")),
        ("Per_Seed_Metrics", read_csv_if_exists(output_dir / "metrics" / "metrics_by_seed.csv")),
        ("Recon_Metrics", read_csv_if_exists(output_dir / "metrics" / "indicator_reconstruction_metrics.csv")),
        ("Event_Window_Stress", read_csv_if_exists(output_dir / "stress_tests" / "event_window_stress_summary.csv")),
        ("Stress107_Setting", read_csv_if_exists(output_dir / "stress_tests" / "stress107_setting.csv")),
        ("Stress107_Window", read_csv_if_exists(output_dir / "stress_tests" / "stress107_window_summary.csv")),
        ("Stress107_Detect", read_csv_if_exists(output_dir / "stress_tests" / "stress107_detection_summary.csv")),
        ("Stress107_Monotonic", read_csv_if_exists(output_dir / "stress_tests" / "stress107_severity_monotonicity.csv")),
        ("Stress107_Key", read_csv_if_exists(output_dir / "stress_tests" / "stress107_key_conclusions.csv")),
        ("CPU_Timing_Summary", read_csv_if_exists(output_dir / "timing" / "cpu_only_inference_timing_summary.csv")),
        ("CPU_Timing_Raw", read_csv_if_exists(output_dir / "timing" / "cpu_only_inference_timing.csv")),
        ("Bootstrap_CI", read_csv_if_exists(output_dir / "stats" / "bootstrap_ci.csv")),
        ("Paired_Tests", read_csv_if_exists(output_dir / "stats" / "paired_error_tests.csv")),
        ("Error_By_WQI_Band", read_csv_if_exists(output_dir / "metrics" / "error_by_wqi_band.csv")),
        ("Prediction_Counts", prediction_counts(output_dir)),
        ("Environment", environment_frame(output_dir)),
        ("Output_File_Index", file_index(output_dir)),
    ]

    wb = Workbook()
    wb.remove(wb.active)
    for name, frame in sheet_specs:
        add_sheet(wb, name, frame)
    ws = wb["README"]
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].fill = PatternFill("solid", fgColor="D9EAF7")
        row[0].font = Font(bold=True)
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(output_file)
    print(output_file)


if __name__ == "__main__":
    main()
